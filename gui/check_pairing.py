import openpyxl

wb = openpyxl.load_workbook(r"G:\drone\Drone-detection-dataset-must-cite\Drone-detection-dataset-master\Data\Video_dataset_description.xlsx")
ws = wb["Blad1"]

# Find AIRPLANE_040 rows for both sensors
print("--- All AIRPLANE entries around 040 ---")
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
    sensor = str(row[0].value) if row[0].value else ""
    cls = str(row[1].value) if row[1].value else ""
    num = row[2].value
    dist = str(row[3].value) if row[3].value else ""
    inter = str(row[4].value) if row[4].value else ""
    source = str(row[5].value) if row[5].value else ""
    if cls == "AIRPLANE" and num and int(num) >= 38 and int(num) <= 42:
        print(f"  {sensor:5s} | {cls} | {num:3d} | dist={dist:10s} | inter={inter} | src={source}")

print()
print("--- Summary: count by sensor ---")
sensors = {}
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    s = str(row[0]) if row[0] else "Unknown"
    sensors[s] = sensors.get(s, 0) + 1
for k, v in sorted(sensors.items()):
    print(f"  {k}: {v} videos")

print()
print("--- Internet-sourced videos (INTER BIN != '' and != 'None') ---")
internet_count = {"V": 0, "IR": 0}
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    sensor = str(row[0]) if row[0] else ""
    inter = row[4]
    if inter and str(inter) not in ("", "None"):
        key = sensor if sensor in internet_count else "?"
        internet_count[key] = internet_count.get(key, 0) + 1
print(f"  V (visible) internet-sourced: {internet_count.get('V', 0)}")
print(f"  IR internet-sourced: {internet_count.get('IR', 0)}")

print()
print("--- Distance bins by sensor ---")
dist_bins = {}
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    sensor = str(row[0]) if row[0] else ""
    dist = str(row[3]) if row[3] else "?"
    key = f"{sensor}_{dist}"
    dist_bins[key] = dist_bins.get(key, 0) + 1
for k, v in sorted(dist_bins.items()):
    print(f"  {k}: {v}")
